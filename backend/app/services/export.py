import csv
import io
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.invoice import Invoice


def _apply_filters(
    query,
    status: str | None,
    date_from: date | None,
    date_to: date | None,
    search: str | None,
):
    if status:
        query = query.where(Invoice.status == status)
    if date_from:
        query = query.where(Invoice.invoice_date >= date_from)
    if date_to:
        query = query.where(Invoice.invoice_date <= date_to)
    if search:
        pattern = f"%{search}%"
        query = query.where(
            or_(
                Invoice.vendor_name.ilike(pattern),
                Invoice.invoice_number.ilike(pattern),
            )
        )
    return query


async def fetch_invoices_for_export(
    db: AsyncSession,
    status: str | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    search: str | None = None,
) -> list[Invoice]:
    query = select(Invoice).order_by(Invoice.created_at.desc())
    query = _apply_filters(query, status, date_from, date_to, search)
    result = await db.execute(query)
    return list(result.scalars().all())


HEADERS = [
    "ID",
    "Status",
    "Invoice Number",
    "Vendor",
    "Invoice Date",
    "Due Date",
    "Total Amount",
    "Tax Amount",
    "Currency",
    "Confidence",
    "Extraction Method",
    "Created At",
]


def _row_from_invoice(inv: Invoice) -> list:
    return [
        str(inv.id),
        inv.status,
        inv.invoice_number or "",
        inv.vendor_name or "",
        inv.invoice_date.isoformat() if inv.invoice_date else "",
        inv.due_date.isoformat() if inv.due_date else "",
        str(inv.total_amount) if inv.total_amount is not None else "",
        str(inv.tax_amount) if inv.tax_amount is not None else "",
        inv.currency or "",
        str(inv.overall_confidence) if inv.overall_confidence is not None else "",
        inv.extraction_method or "",
        inv.created_at.isoformat() if inv.created_at else "",
    ]


def generate_csv(invoices: list[Invoice]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(HEADERS)
    for inv in invoices:
        writer.writerow(_row_from_invoice(inv))
    return output.getvalue().encode("utf-8")


def generate_excel(invoices: list[Invoice]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for row_idx, inv in enumerate(invoices, 2):
        for col_idx, value in enumerate(_row_from_invoice(inv), 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = "A2"

    for col in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col)
        max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[letter].width = min(max_len + 2, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
